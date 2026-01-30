import os
import re
import time
import asyncio
from pathlib import Path
from datetime import datetime
from telethon import TelegramClient, events, Button
from telethon.tl.types import DocumentAttributeFilename
from telethon.tl.functions.channels import CreateChannelRequest
import py7zr
import zipfile
import tarfile
import gzip
import shutil
import config
from concurrent.futures import ThreadPoolExecutor
import threading
import csv
import openpyxl

# Initialize bot for handling commands
bot = TelegramClient('bot_session', config.API_ID, config.API_HASH).start(bot_token=config.BOT_TOKEN)

# Initialize user client for downloading from channel (will prompt for phone/code on first run)
user_client = TelegramClient('user_session', config.API_ID, config.API_HASH)

# Cache directory for temporary extracted files
CACHE_DIR = Path("./cache")
CACHE_DIR.mkdir(exist_ok=True)

# Tracking file for processed archives
PROCESSED_FILES_TRACK = Path("./processed_archives.txt")

# Global variables
extracted_channel_id = config.EXTRACTED_CHANNEL_ID
extracted_files_messages = {}  # Map filename to message ID in extracted channel
cache_loaded = False
data_cache = []  # In-memory cache for all data (used only if ENABLE_FULL_CACHE = True)
cache_lock = threading.Lock()
cache_size_mb = 0  # Track cache size

# Download control
download_process_running = False
download_process_paused = True  # Default: STOPPED (manual start required)
download_task = None
stats = {
    'total_files': 0,
    'loaded_files': 0,
    'total_size_mb': 0,
    'total_lines': 0,
    'processing_file': None
}


def load_processed_files():
    """Load list of already processed .7z files"""
    if PROCESSED_FILES_TRACK.exists():
        with open(PROCESSED_FILES_TRACK, 'r', encoding='utf-8') as f:
            return set(line.strip() for line in f if line.strip())
    return set()


def mark_file_as_processed(filename):
    """Mark a .7z file as processed"""
    with open(PROCESSED_FILES_TRACK, 'a', encoding='utf-8') as f:
        f.write(f"{filename}\n")


def extract_archive_sync(archive_path, extract_dir):
    """Synchronous extraction for threading - supports multiple formats"""
    try:
        filename = archive_path.name.lower()
        
        # 7z files
        if filename.endswith('.7z'):
            with py7zr.SevenZipFile(archive_path, mode='r') as archive:
                archive.extractall(path=extract_dir)
        
        # ZIP files
        elif filename.endswith('.zip'):
            with zipfile.ZipFile(archive_path, 'r') as archive:
                archive.extractall(path=extract_dir)
        
        # RAR files (note: requires rarfile library)
        elif filename.endswith('.rar'):
            try:
                import rarfile
                with rarfile.RarFile(archive_path, 'r') as archive:
                    archive.extractall(path=extract_dir)
            except ImportError:
                print("Warning: rarfile not installed. Install with: pip install rarfile")
                return False
        
        # TAR.GZ / TGZ files
        elif filename.endswith(('.tar.gz', '.tgz')):
            with tarfile.open(archive_path, 'r:gz') as archive:
                archive.extractall(path=extract_dir)
        
        # TAR files
        elif filename.endswith('.tar'):
            with tarfile.open(archive_path, 'r') as archive:
                archive.extractall(path=extract_dir)
        
        # GZ files (single file compression)
        elif filename.endswith('.gz') and not filename.endswith('.tar.gz'):
            output_file = extract_dir / archive_path.stem
            extract_dir.mkdir(parents=True, exist_ok=True)
            with gzip.open(archive_path, 'rb') as f_in:
                with open(output_file, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
        
        # BZ2 files
        elif filename.endswith('.bz2'):
            import bz2
            output_file = extract_dir / archive_path.stem
            extract_dir.mkdir(parents=True, exist_ok=True)
            with bz2.open(archive_path, 'rb') as f_in:
                with open(output_file, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
        
        # XZ files
        elif filename.endswith('.xz'):
            import lzma
            output_file = extract_dir / archive_path.stem
            extract_dir.mkdir(parents=True, exist_ok=True)
            with lzma.open(archive_path, 'rb') as f_in:
                with open(output_file, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
        
        else:
            print(f"Unsupported archive format: {filename}")
            return False
        
        return True
        
    except Exception as e:
        print(f"Error extracting {archive_path.name}: {e}")
        return False

# Search patterns for different data types
PATTERNS = {
    'domain': re.compile(r'(?:https?://)?([a-zA-Z0-9-]+(?:\.[a-zA-Z0-9-]+)+)', re.IGNORECASE),
    'email': re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', re.IGNORECASE),
    'num': re.compile(r'\b\d{10,15}\b'),  # Phone numbers
    'aadhar': re.compile(r'\b\d{4}\s?\d{4}\s?\d{4}\b'),  # Aadhar format
    'pan': re.compile(r'\b[A-Z]{5}\d{4}[A-Z]\b'),  # PAN card format
    'ps': re.compile(r':([^:\n]+)$', re.MULTILINE),  # Password (last field after colon)
    'tg': re.compile(r'(?:t\.me|telegram\.(?:me|org))/([a-zA-Z0-9_]+)', re.IGNORECASE),
    'ig': re.compile(r'(?:instagram\.com|instagr\.am)/([a-zA-Z0-9._]+)', re.IGNORECASE),
    'fb': re.compile(r'(?:facebook\.com|fb\.com)/([a-zA-Z0-9.]+)', re.IGNORECASE),
    'lk': re.compile(r'linkedin\.com/in/([a-zA-Z0-9-]+)', re.IGNORECASE),
    'x': re.compile(r'(?:twitter\.com|x\.com)/([a-zA-Z0-9_]+)', re.IGNORECASE),
    'yt': re.compile(r'youtube\.com/(?:c/|channel/|@)?([a-zA-Z0-9_-]+)', re.IGNORECASE),
    'wa': re.compile(r'(?:wa\.me|whatsapp\.com)/(\d+)', re.IGNORECASE),
    'bi': re.compile(r'binance', re.IGNORECASE),
    'ot': None  # General text search
}


async def notify_admins(message):
    """Send notification to all admins"""
    for admin_id in config.ADMIN_IDS:
        try:
            await bot.send_message(admin_id, message)
        except Exception as e:
            print(f"Error notifying admin {admin_id}: {e}")


async def verify_extracted_channel():
    """Verify extracted channel exists and is accessible"""
    global extracted_channel_id
    
    if not extracted_channel_id:
        print("❌ No extracted channel configured!")
        await notify_admins("❌ No extracted channel configured in config.py!")
        return False
    
    try:
        await user_client.start()
        channel = await user_client.get_entity(extracted_channel_id)
        print(f"✅ Using existing channel: {channel.title} (ID: {extracted_channel_id})")
        await notify_admins(f"✅ Using existing channel: {channel.title}")
        return True
        
    except Exception as e:
        print(f"❌ Cannot access extracted channel: {e}")
        await notify_admins(f"❌ Cannot access extracted channel {extracted_channel_id}: {e}")
        return False


def is_text_file(file_path):
    """Check if a file is text-based by trying to read it"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='strict') as f:
            f.read(1024)  # Try reading first 1KB
        return True
    except (UnicodeDecodeError, Exception):
        # Try with different encodings
        try:
            with open(file_path, 'r', encoding='latin-1', errors='ignore') as f:
                f.read(1024)
            return True
        except:
            return False


async def load_data_into_memory():
    """Load all extracted text-based files into memory for ultra-fast searching"""
    global data_cache, cache_loaded
    
    print("📥 Loading all data into memory for fast search...")
    data_cache = []
    
    try:
        # Get ALL files from cache directory (not just .txt)
        all_files = [f for f in CACHE_DIR.rglob('*') if f.is_file()]
        total_files = len(all_files)
        loaded = 0
        skipped = 0
        
        print(f"🔍 Found {total_files} files to process")
        
        # Common text-based extensions
        text_extensions = {
            '.txt', '.log', '.csv', '.json', '.xml', '.sql', '.md', 
            '.html', '.htm', '.css', '.js', '.py', '.java', '.c', 
            '.cpp', '.h', '.php', '.rb', '.go', '.rs', '.sh', '.bat',
            '.ini', '.conf', '.cfg', '.yaml', '.yml', '.toml', '.env',
            '.dat', '.list', '.lst', '.dump', '.leak'
        }
        
        for file_path in all_files:
            try:
                # Check extension first (fast check)
                if file_path.suffix.lower() in text_extensions:
                    # Known text file
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        data_cache.extend(f.readlines())
                    loaded += 1
                elif file_path.suffix == '' or file_path.suffix not in ['.zip', '.7z', '.rar', '.tar', '.gz', '.exe', '.dll', '.bin', '.img', '.iso']:
                    # Unknown extension or no extension - try to read as text
                    if is_text_file(file_path):
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            data_cache.extend(f.readlines())
                        loaded += 1
                    else:
                        skipped += 1
                else:
                    # Binary file extension
                    skipped += 1
                
                if (loaded + skipped) % 20 == 0:
                    print(f"Progress: {loaded} loaded, {skipped} skipped ({loaded + skipped}/{total_files})")
                    
            except Exception as e:
                print(f"Error loading {file_path.name}: {e}")
                skipped += 1
        
        cache_loaded = True
        print(f"✅ Loaded {len(data_cache):,} lines from {loaded} files")
        print(f"⏭️ Skipped {skipped} binary/non-text files")
        
        await notify_admins(
            f"✅ **Memory Cache Ready**\n\n"
            f"📊 Loaded: {len(data_cache):,} lines from {loaded} files\n"
            f"⏭️ Skipped: {skipped} binary files\n"
            f"Bot is now ready for ultra-fast queries!"
        )
        
    except Exception as e:
        print(f"Error loading data into memory: {e}")
        await notify_admins(f"❌ Error loading data: {e}")


async def upload_extracted_files():
    """Upload all extracted files to the channel (all file types)"""
    global extracted_files_messages
    
    try:
        await user_client.start()
        
        # Get or create extracted files channel
        if extracted_channel_id is None:
            channel_id = await create_extracted_channel()
            if channel_id is None:
                return
        else:
            channel_id = extracted_channel_id
        
        channel = await user_client.get_entity(channel_id)
        
        # Find ALL files (not just .txt)
        all_files = [f for f in CACHE_DIR.rglob('*') if f.is_file() and not f.name.endswith('.7z')]
        total_files = len(all_files)
        
        print(f"Found {total_files} extracted files to upload (all types)")
        await notify_admins(f"📤 Uploading {total_files} extracted files to channel (all types)...")
        
        uploaded = 0
        
        # Upload all files
        for file_path in all_files:
            try:
                # Determine emoji based on file type
                ext = file_path.suffix.lower()
                emoji = "📄"
                if ext in ['.txt', '.log', '.dat', '.list', '.dump', '.leak']:
                    emoji = "📝"
                elif ext in ['.csv', '.sql', '.json', '.xml']:
                    emoji = "📊"
                elif ext in ['.zip', '.rar', '.tar', '.gz']:
                    emoji = "📦"
                
                # Upload file to channel with maximum speed
                message = await user_client.send_file(
                    channel,
                    file_path,
                    caption=f"{emoji} {file_path.name}",
                    force_document=True,
                    part_size_kb=512  # Use larger chunks for faster upload
                )
                
                extracted_files_messages[file_path.name] = message.id
                uploaded += 1
                
                if uploaded % 10 == 0:  # Notify every 10 files
                    print(f"Uploaded {uploaded}/{total_files} files")
                    await notify_admins(f"📤 Progress: {uploaded}/{total_files} files uploaded")
                
            except Exception as e:
                print(f"Error uploading {file_path.name}: {e}")
        
        print(f"✅ Successfully uploaded {uploaded} files")
        await notify_admins(
            f"✅ **Upload Complete**\n\n"
            f"Successfully uploaded {uploaded}/{total_files} files to channel\n"
            f"(All file types included)"
        )
        
        # Now load everything into memory for fast search
        await load_data_into_memory()
        
    except Exception as e:
        print(f"Error in upload_extracted_files: {e}")
        await notify_admins(f"❌ Error uploading files: {e}")


async def download_and_extract_archives():
    """Download all .7z files from the channel and extract them locally with parallel processing"""
    global cache_loaded, download_process_running, stats
    
    download_process_running = True
    
    # Wait if paused (only log once)
    if download_process_paused:
        print("⏸️ Process is PAUSED. Waiting for manual START...")
        while download_process_paused:
            await asyncio.sleep(2)
    
    print("⚡ Starting FAST download and extraction from source channel...")
    
    # Load already processed files
    processed_files = load_processed_files()
    print(f"Already processed {len(processed_files)} archives")
    
    try:
        # Start user client
        await user_client.start()
        
        # Get the source channel entity
        channel = await user_client.get_entity(config.SOURCE_CHANNEL_ID)
        
        # Collect all .7z files that need processing
        files_to_process = []
        
        print("🔍 Scanning channel for all files (oldest to newest)...")
        
        # Get all messages first, then reverse to process oldest first
        all_messages = []
        async for message in user_client.iter_messages(channel, limit=None):
            all_messages.append(message)
        
        # Reverse to get oldest first
        all_messages.reverse()
        
        # Supported archive extensions
        archive_extensions = ('.7z', '.zip', '.rar', '.tar', '.gz', '.tar.gz', '.tgz', '.bz2', '.xz')
        
        # Also check for direct data files (xlsx, txt, csv, sql)
        data_file_extensions = ('.xlsx', '.xls', '.txt', '.csv', '.sql', '.json', '.xml', '.log', '.dat')
        
        for message in all_messages:
            if message.document:
                filename = None
                for attr in message.document.attributes:
                    if isinstance(attr, DocumentAttributeFilename):
                        filename = attr.file_name
                        break
                
                if filename:
                    # Check if it's any supported archive type OR data file
                    if filename.lower().endswith(archive_extensions + data_file_extensions):
                        if filename not in processed_files:
                            files_to_process.append((message, filename))
                        else:
                            print(f"⏭️ Skipping: {filename}")
        
        if not files_to_process:
            print("✅ No new archives to process")
            await notify_admins("✅ No new archives found. All files already processed!")
            
            # Mark as ready if channel exists
            if extracted_channel_id:
                cache_loaded = True
                print("✅ Bot ready! Channel files available for search.")
                await notify_admins("✅ Bot ready! You can start searching now.")
            return
        
        print(f"🚀 Found {len(files_to_process)} new archives to process")
        stats['total_files'] = len(files_to_process)
        await notify_admins(f"🚀 Found {len(files_to_process)} new archives. Starting parallel processing...")
        
        # Process files with better progress tracking
        total = len(files_to_process)
        
        # Use ThreadPoolExecutor for parallel extraction (increase workers for speed)
        executor = ThreadPoolExecutor(max_workers=4)  # Increased from 2 to 4
        
        # Track progress message for live updates
        progress_msg_id = {}
        
        for idx, (message, filename) in enumerate(files_to_process, 1):
            # Check if paused
            if download_process_paused:
                print("⏸️ Process PAUSED by admin")
                await notify_admins("⏸️ Download process PAUSED by admin")
                download_process_running = False
                return
            
            stats['processing_file'] = filename
            file_path = CACHE_DIR / filename
            
            # Check if it's a data file (not archive)
            data_file_extensions = ('.txt', '.csv', '.sql', '.json', '.xml', '.log', '.dat')
            excel_extensions = ('.xlsx', '.xls', '.xlsm')
            is_data_file = filename.lower().endswith(data_file_extensions)
            is_excel_file = filename.lower().endswith(excel_extensions)
            
            # For archives: prepare extract directory
            if not is_data_file:
                extract_name = filename
                for ext in ['.7z', '.zip', '.rar', '.tar.gz', '.tgz', '.tar', '.gz', '.bz2', '.xz']:
                    if extract_name.endswith(ext):
                        extract_name = extract_name[:-len(ext)]
                        break
                extract_dir = CACHE_DIR / extract_name
            else:
                extract_dir = None
            
            # Download with real-time progress
            print(f"⬇️ [{idx}/{total}] Downloading: {filename}")
            
            # Send initial notification to admins
            for admin_id in config.ADMIN_IDS:
                try:
                    msg = await bot.send_message(
                        admin_id, 
                        f"⬇️ **[{idx}/{total}] Downloading**\n\n"
                        f"📦 File: `{filename}`\n"
                        f"📊 Progress: 0%\n"
                        f"⚡ Speed: Calculating..."
                    )
                    progress_msg_id[admin_id] = msg.id
                except:
                    pass
            
            last_update_time = [time.time()]
            last_percentage = [0]
            download_start_time = time.time()
            
            async def progress_callback(current, total_size):
                """Show download progress with live updates to admin"""
                if total_size > 0:
                    percentage = int((current / total_size) * 100)
                    current_mb = current / (1024 * 1024)
                    total_mb = total_size / (1024 * 1024)
                    elapsed = time.time() - download_start_time
                    speed_mbps = (current_mb / elapsed) if elapsed > 0 else 0
                    
                    # Update every 0.3 seconds for smooth animation
                    now = time.time()
                    if now - last_update_time[0] >= 0.3:
                        # Console output
                        print(f"   📥 Progress: {percentage}% ({current_mb:.1f}/{total_mb:.1f} MB) - {speed_mbps:.2f} MB/s")
                        
                        # Update admin messages
                        progress_bar = "█" * (percentage // 5) + "░" * (20 - percentage // 5)
                        for admin_id in config.ADMIN_IDS:
                            try:
                                if admin_id in progress_msg_id:
                                    await bot.edit_message(
                                        admin_id,
                                        progress_msg_id[admin_id],
                                        f"⬇️ **[{idx}/{total}] Downloading**\n\n"
                                        f"📦 File: `{filename}`\n"
                                        f"📊 Progress: {percentage}%\n"
                                        f"[{progress_bar}]\n"
                                        f"💾 {current_mb:.1f} / {total_mb:.1f} MB\n"
                                        f"⚡ Speed: {speed_mbps:.2f} MB/s"
                                    )
                            except:
                                pass
                        
                        last_percentage[0] = percentage
                        last_update_time[0] = now
            
            start_time = time.time()
            try:
                # Download with maximum speed settings
                await user_client.download_media(
                    message, 
                    file=str(file_path),
                    progress_callback=progress_callback
                )
                
                download_time = time.time() - start_time
                size_mb = file_path.stat().st_size / (1024 * 1024)
                speed_mbps = (size_mb / download_time) if download_time > 0 else 0
                stats['total_size_mb'] += size_mb
                print(f"✅ Downloaded {filename} ({size_mb:.1f} MB) in {download_time:.1f}s ({speed_mbps:.2f} MB/s)")
                
            except Exception as e:
                print(f"❌ Download failed for {filename}: {e}")
                await notify_admins(f"❌ Download failed for {filename}: {e}\nRetrying...")
                
                # Retry once
                print(f"🔄 Retrying download: {filename}")
                try:
                    await user_client.download_media(
                        message, 
                        file=str(file_path),
                        progress_callback=progress_callback
                    )
                    print(f"✅ Downloaded {filename} on retry")
                except Exception as retry_error:
                    print(f"❌ Retry failed: {retry_error}")
                    await notify_admins(f"❌ Could not download {filename} after retry. Skipping...")
                    continue
            
            # If it's an Excel file, convert to CSV first
            if is_excel_file:
                print(f"📊 Excel file detected: {filename} (converting to CSV)")
                
                # Update admin - converting
                for admin_id in config.ADMIN_IDS:
                    try:
                        if admin_id in progress_msg_id:
                            await bot.edit_message(
                                admin_id,
                                progress_msg_id[admin_id],
                                f"✅ **Download Complete!**\n\n"
                                f"📊 File: `{filename}`\n"
                                f"💾 Size: {size_mb:.1f} MB\n"
                                f"⚡ Speed: {speed_mbps:.2f} MB/s\n\n"
                                f"🔄 Converting Excel to CSV..."
                            )
                    except:
                        pass
                
                try:
                    # Convert Excel to CSV
                    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    converted_lines = []
                    
                    for sheet_name in workbook.sheetnames:
                        print(f"   Reading sheet: {sheet_name}")
                        sheet = workbook[sheet_name]
                        
                        for row in sheet.iter_rows(values_only=True):
                            row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                            if row_text.strip():
                                converted_lines.append(row_text + '\n')
                    
                    workbook.close()
                    
                    # Save as CSV
                    csv_filename = filename.rsplit('.', 1)[0] + '.csv'
                    csv_path = CACHE_DIR / csv_filename
                    
                    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
                        f.writelines(converted_lines)
                    
                    print(f"✅ Converted {len(converted_lines):,} rows to {csv_filename}")
                    
                    # Update admin - uploading CSV
                    for admin_id in config.ADMIN_IDS:
                        try:
                            if admin_id in progress_msg_id:
                                await bot.edit_message(
                                    admin_id,
                                    progress_msg_id[admin_id],
                                    f"✅ **Conversion Complete!**\n\n"
                                    f"📊 Original: `{filename}`\n"
                                    f"📄 CSV: `{csv_filename}`\n"
                                    f"📝 Rows: {len(converted_lines):,}\n\n"
                                    f"📤 Uploading CSV to channel..."
                                )
                        except:
                            pass
                    
                    # Upload CSV to channel
                    if not await verify_extracted_channel():
                        print("❌ Cannot upload - channel not accessible")
                        mark_file_as_processed(filename)
                        stats['loaded_files'] += 1
                        continue
                    
                    channel_id = extracted_channel_id
                    channel = await user_client.get_entity(channel_id)
                    
                    await user_client.send_file(
                        channel,
                        csv_path,
                        caption=f"📊 {csv_filename} (auto-converted from {filename})",
                        force_document=True,
                        part_size_kb=512
                    )
                    print(f"✅ Uploaded {csv_filename} to channel")
                    
                    # Mark cache as loaded
                    cache_loaded = True
                    
                    # Update admin - complete
                    for admin_id in config.ADMIN_IDS:
                        try:
                            if admin_id in progress_msg_id:
                                await bot.edit_message(
                                    admin_id,
                                    progress_msg_id[admin_id],
                                    f"✅ **Complete!**\n\n"
                                    f"📊 Excel: `{filename}`\n"
                                    f"📄 CSV: `{csv_filename}`\n"
                                    f"💾 Rows: {len(converted_lines):,}\n\n"
                                    f"✅ Uploaded to channel!"
                                )
                        except:
                            pass
                    
                    # Load into memory
                    with cache_lock:
                        data_cache.extend(converted_lines)
                    stats['total_lines'] = len(data_cache)
                    print(f"💾 Loaded {len(converted_lines):,} lines into memory cache")
                    
                    # Clean up
                    file_path.unlink()  # Delete Excel
                    csv_path.unlink()   # Delete CSV (already uploaded)
                    
                except Exception as e:
                    print(f"❌ Error converting Excel: {e}")
                    await notify_admins(f"❌ Error converting {filename}: {e}")
                
                mark_file_as_processed(filename)
                stats['loaded_files'] += 1
                continue
            
            # If it's a data file (not archive), upload immediately
            if is_data_file:
                print(f"📄 Data file detected: {filename} (no extraction needed)")
                
                # Update admin - uploading to channel
                for admin_id in config.ADMIN_IDS:
                    try:
                        if admin_id in progress_msg_id:
                            await bot.edit_message(
                                admin_id,
                                progress_msg_id[admin_id],
                                f"✅ **Download Complete!**\n\n"
                                f"📄 File: `{filename}`\n"
                                f"💾 Size: {size_mb:.1f} MB\n"
                                f"⚡ Speed: {speed_mbps:.2f} MB/s\n\n"
                                f"📤 Uploading to channel..."
                            )
                    except:
                        pass
                
                # Use existing extracted files channel
                if not await verify_extracted_channel():
                    print("❌ Cannot upload - channel not accessible")
                    mark_file_as_processed(filename)
                    stats['loaded_files'] += 1
                    continue
                
                channel_id = extracted_channel_id
                
                # Upload to channel
                try:
                    channel = await user_client.get_entity(channel_id)
                    emoji = "📄"
                    if filename.endswith('.csv'):
                        emoji = "📊"
                    elif filename.endswith('.sql'):
                        emoji = "💾"
                    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                        emoji = "📈"
                    
                    await user_client.send_file(
                        channel,
                        file_path,
                        caption=f"{emoji} {filename}",
                        force_document=True
                    )
                    print(f"✅ Uploaded {filename} to channel")
                    
                    # Mark cache as loaded so searches can start
                    cache_loaded = True
                    
                    # Update admin - complete
                    for admin_id in config.ADMIN_IDS:
                        try:
                            if admin_id in progress_msg_id:
                                await bot.edit_message(
                                    admin_id,
                                    progress_msg_id[admin_id],
                                    f"✅ **Complete!**\n\n"
                                    f"📄 File: `{filename}`\n"
                                    f"💾 Size: {size_mb:.1f} MB\n"
                                    f"⚡ Speed: {speed_mbps:.2f} MB/s\n\n"
                                    f"✅ Uploaded to channel!"
                                )
                        except:
                            pass
                    
                    # Load into memory for search
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            new_lines = f.readlines()
                            with cache_lock:
                                data_cache.extend(new_lines)
                            print(f"💾 Loaded {len(new_lines)} lines into memory cache")
                    except:
                        print(f"⚠️ Could not load {filename} into memory (might be binary)")
                    
                except Exception as e:
                    print(f"❌ Error uploading {filename}: {e}")
                    await notify_admins(f"❌ Error uploading {filename}: {e}")
                
                mark_file_as_processed(filename)
                stats['loaded_files'] += 1
                continue
            
            # Update admin messages - download complete, starting extraction
            for admin_id in config.ADMIN_IDS:
                try:
                    if admin_id in progress_msg_id:
                        await bot.edit_message(
                            admin_id,
                            progress_msg_id[admin_id],
                            f"✅ **Download Complete!**\n\n"
                            f"📦 File: `{filename}`\n"
                            f"💾 Size: {size_mb:.1f} MB\n"
                            f"⚡ Speed: {speed_mbps:.2f} MB/s\n\n"
                            f"📦 Extracting now..."
                        )
                except:
                    pass
            
            # Extract in thread pool (non-blocking)
            print(f"📦 [{idx}/{total}] Extracting: {filename}")
            
            extract_start = time.time()
            loop = asyncio.get_event_loop()
            success = await loop.run_in_executor(
                executor, 
                extract_archive_sync, 
                file_path, 
                extract_dir
            )
            
            extract_time = time.time() - extract_start
            
            if success:
                print(f"✅ Extracted {filename} in {extract_time:.1f}s")
                
                # Upload extracted files to channel immediately
                print(f"📤 Uploading extracted files from {filename} to channel...")
                
                # Use existing extracted files channel
                if not await verify_extracted_channel():
                    print("❌ Cannot upload - channel not accessible")
                    mark_file_as_processed(filename)
                    stats['loaded_files'] += 1
                    continue
                
                channel_id = extracted_channel_id
                channel = await user_client.get_entity(channel_id)
                
                # Process extracted files - convert Excel to CSV
                all_extracted_files = [f for f in extract_dir.rglob('*') if f.is_file()]
                print(f"   📦 Found {len(all_extracted_files)} extracted files")
                
                files_to_upload = []
                excel_converted = 0
                
                for extracted_file in all_extracted_files:
                    # Check if it's an Excel file
                    if extracted_file.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
                        print(f"      🔄 Converting Excel: {extracted_file.name}")
                        
                        try:
                            # Convert to CSV
                            workbook = openpyxl.load_workbook(extracted_file, read_only=True, data_only=True)
                            converted_lines = []
                            
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                for row in sheet.iter_rows(values_only=True):
                                    row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                                    if row_text.strip():
                                        converted_lines.append(row_text + '\n')
                            
                            workbook.close()
                            
                            # Save as CSV
                            csv_filename = extracted_file.stem + '.csv'
                            csv_path = extracted_file.parent / csv_filename
                            
                            with open(csv_path, 'w', encoding='utf-8', newline='') as f:
                                f.writelines(converted_lines)
                            
                            print(f"         ✅ Converted to {csv_filename} ({len(converted_lines):,} rows)")
                            
                            # Add CSV to upload list (not the Excel file)
                            files_to_upload.append(csv_path)
                            excel_converted += 1
                            
                            # Delete the Excel file
                            extracted_file.unlink()
                            
                        except Exception as e:
                            print(f"         ❌ Error converting {extracted_file.name}: {e}")
                            # If conversion fails, upload the Excel file anyway
                            files_to_upload.append(extracted_file)
                    else:
                        # Not Excel - add to upload list
                        files_to_upload.append(extracted_file)
                
                if excel_converted > 0:
                    print(f"   ✅ Converted {excel_converted} Excel files to CSV")
                
                # Upload all files
                uploaded_count = 0
                
                for file_to_upload in files_to_upload:
                    try:
                        # Determine emoji
                        ext = file_to_upload.suffix.lower()
                        emoji = "📄"
                        if ext in ['.txt', '.log', '.dat', '.list', '.dump', '.leak']:
                            emoji = "📝"
                        elif ext in ['.csv', '.sql', '.json', '.xml']:
                            emoji = "📊"
                        
                        await user_client.send_file(
                            channel,
                            file_to_upload,
                            caption=f"{emoji} {file_to_upload.name}",
                            force_document=True,
                            part_size_kb=512
                        )
                        
                        uploaded_count += 1
                        
                        # Load text files into memory cache
                        if is_text_file(file_to_upload):
                            try:
                                with open(file_to_upload, 'r', encoding='utf-8', errors='ignore') as f:
                                    new_lines = f.readlines()
                                    with cache_lock:
                                        data_cache.extend(new_lines)
                            except:
                                pass
                        
                    except Exception as e:
                        print(f"⚠️ Error uploading {file_to_upload.name}: {e}")
                
                print(f"✅ Uploaded {uploaded_count} files from {filename} to channel")
                
                # Mark cache as loaded so searches can start
                cache_loaded = True
                
                # Update admin - complete
                for admin_id in config.ADMIN_IDS:
                    try:
                        if admin_id in progress_msg_id:
                            await bot.edit_message(
                                admin_id,
                                progress_msg_id[admin_id],
                                f"✅ **Complete!**\n\n"
                                f"📦 Archive: `{filename}`\n"
                                f"📤 Uploaded: {uploaded_count} files\n"
                                f"⏱️ Extract time: {extract_time:.1f}s\n\n"
                                f"✅ Ready for search!"
                            )
                    except:
                        pass
                
                mark_file_as_processed(filename)
                cache_loaded = True  # Mark cache as loaded
                stats['loaded_files'] += 1
                
                # Delete archive to save space
                file_path.unlink()
                print(f"🗑️ Deleted archive: {filename}")
                
                # Clean up extracted directory
                shutil.rmtree(extract_dir)
                print(f"🗑️ Cleaned up extracted directory")
                
            else:
                await notify_admins(f"❌ Error extracting {filename}")
                continue
        
        executor.shutdown(wait=True)
        
        # All files have been uploaded individually during processing
        total_lines = len(data_cache)
        stats['total_lines'] = total_lines
        stats['processing_file'] = None
        download_process_running = False
        
        await notify_admins(
            f"✅ **All Processing Complete!**\n\n"
            f"📊 Total files processed: {len(files_to_process)}\n"
            f"💾 Total data lines loaded: {total_lines:,}\n\n"
            f"🚀 Bot is ready for ultra-fast queries!"
        )
        
    except Exception as e:
        print(f"❌ Error in download_and_extract_archives: {e}")
        await notify_admins(f"❌ Error in download/extract: {e}")


async def search_data(query, search_type):
    """Search through channel files - memory efficient version"""
    results = []
    query_lower = query.lower()
    
    # If full cache is enabled and loaded, use it (fastest)
    if config.ENABLE_FULL_CACHE and len(data_cache) > 0:
        print(f"🔍 Searching in-memory cache ({len(data_cache):,} lines)...")
        print(f"🔍 Query: '{query_lower}'")
        
        # Debug: Show sample of first and last few lines
        if len(data_cache) > 0:
            print(f"📝 First 5 lines in cache:")
            for i, sample_line in enumerate(data_cache[:5]):
                # Show more characters and check encoding
                display = sample_line[:200].replace('\n', '\\n').replace('\r', '\\r').replace('\t', '\\t')
                print(f"   [{i+1}] {display}")
            
            print(f"📝 Last 3 lines in cache:")
            for i, sample_line in enumerate(data_cache[-3:]):
                display = sample_line[:200].replace('\n', '\\n').replace('\r', '\\r').replace('\t', '\\t')
                print(f"   [{len(data_cache)-3+i+1}] {display}")
            
            # Test search on first 10 lines manually
            print(f"🔍 Testing query '{query_lower}' on first 10 lines:")
            for i, line in enumerate(data_cache[:10]):
                if query_lower in line.lower():
                    print(f"   ✅ MATCH on line {i+1}: {line[:100]}")
        
        with cache_lock:
            for line in data_cache:
                if query_lower in line.lower():
                    results.append(line.strip())
        
        print(f"✅ Found {len(results)} matches")
        return results
    
    # Memory-efficient mode: Search directly from channel files (on-demand)
    if extracted_channel_id:
        print(f"🔍 Searching channel files...")
        try:
            channel = await user_client.get_entity(extracted_channel_id)
            
            # Search through all uploaded files in the channel
            async for message in user_client.iter_messages(channel):
                if message.document:
                    try:
                        # Download file to memory
                        file_bytes = await user_client.download_media(message, file=bytes)
                        
                        # Try to decode as text with proper encoding detection
                        file_content = None
                        
                        # Check for BOM
                        if file_bytes[:2] == b'\xff\xfe' or file_bytes[:2] == b'\xfe\xff':
                            file_content = file_bytes.decode('utf-16', errors='ignore')
                        elif file_bytes[:3] == b'\xef\xbb\xbf':
                            file_content = file_bytes[3:].decode('utf-8', errors='ignore')
                        else:
                            # Try common encodings
                            for encoding in ['utf-8', 'utf-16-le', 'utf-16-be', 'latin-1', 'cp1252']:
                                try:
                                    file_content = file_bytes.decode(encoding, errors='strict')
                                    break
                                except:
                                    continue
                        
                        # Fallback
                        if not file_content:
                            file_content = file_bytes.decode('utf-8', errors='ignore')
                        
                        # Search through lines
                        for line in file_content.splitlines():
                            if query_lower in line.lower():
                                results.append(line.strip())
                    except Exception as e:
                        continue
            
            return results
        except Exception as e:
            print(f"Error searching channel: {e}")
            return []
    
    return []


async def get_admin_panel_text():
    """Generate admin panel caption text"""
    status = "🟢 RUNNING" if download_process_running and not download_process_paused else "🔴 STOPPED"
    
    # Get channel info
    source_info = f"`{config.SOURCE_CHANNEL_ID}`"
    extract_info = f"`{extracted_channel_id}`" if extracted_channel_id else "Not created yet"
    
    # Calculate total size
    total_size = stats['total_size_mb']
    
    text = f"""
🔧 **ADMIN CONTROL PANEL**

📊 **System Status**
• Status: {status}
• Cache Loaded: {'✅ Yes' if cache_loaded else '❌ No'}

📂 **Channel Information**
• Source Channel: {source_info}
• Extract Channel: {extract_info}

📈 **Statistics**
• Total Files: {stats['total_files']}
• Loaded Files: {stats['loaded_files']}/{stats['total_files']}
• Total Size: {total_size:.2f} MB
• Total Lines: {stats['total_lines']:,}
• Currently Processing: {stats['processing_file'] or 'None'}

📋 **Available Commands for Users**
/domain - Domain analysis
/email - Email lookup
/num - Number lookup
/aadhar - Aadhar lookup
/pan - PAN lookup
/ps - Password lookup
/tg - Telegram lookup
/ig - Instagram lookup
/fb - Facebook lookup
/lk - LinkedIn lookup
/x - Twitter/X lookup
/yt - YouTube lookup
/wa - WhatsApp lookup
/bi - Binance lookup
/ot - General text lookup
"""
    return text


@bot.on(events.NewMessage(pattern='/admin'))
async def admin_handler(event):
    """Admin panel - only for admins"""
    if event.sender_id not in config.ADMIN_IDS:
        await event.respond("❌ Unauthorized. Admin only!")
        return
    
    text = await get_admin_panel_text()
    
    # Create inline buttons
    button_text = "🟢 STOP" if (download_process_running and not download_process_paused) else "🔴 START"
    
    buttons = [
        [Button.inline(button_text, b"toggle_process")],
        [Button.inline("🔄 Refresh Stats", b"refresh_stats")],
        [Button.inline("📊 Channel Stats", b"channel_stats")],
        [Button.inline("🗑️ Clear Cache", b"clear_cache")]
    ]
    
    await event.respond(text, buttons=buttons)


@bot.on(events.CallbackQuery(pattern=b"toggle_process"))
async def toggle_process_handler(event):
    """Toggle download process start/stop"""
    global download_process_paused, download_task
    
    if event.sender_id not in config.ADMIN_IDS:
        await event.answer("❌ Unauthorized!", alert=True)
        return
    
    download_process_paused = not download_process_paused
    
    if not download_process_paused:
        # START process
        await event.answer("🟢 Starting download process...")
        if not download_process_running:
            download_task = asyncio.create_task(download_and_extract_archives())
    else:
        # STOP process
        await event.answer("🔴 Stopping download process...")
    
    # Update panel
    text = await get_admin_panel_text()
    button_text = "🟢 STOP" if not download_process_paused else "🔴 START"
    
    buttons = [
        [Button.inline(button_text, b"toggle_process")],
        [Button.inline("🔄 Refresh Stats", b"refresh_stats")],
        [Button.inline("📊 Channel Stats", b"channel_stats")],
        [Button.inline("🗑️ Clear Cache", b"clear_cache")]
    ]
    
    await event.edit(text, buttons=buttons)


@bot.on(events.CallbackQuery(pattern=b"refresh_stats"))
async def refresh_stats_handler(event):
    """Refresh statistics display"""
    if event.sender_id not in config.ADMIN_IDS:
        await event.answer("❌ Unauthorized!", alert=True)
        return
    
    await event.answer("🔄 Refreshing...")
    
    text = await get_admin_panel_text()
    button_text = "🟢 STOP" if (download_process_running and not download_process_paused) else "🔴 START"
    
    buttons = [
        [Button.inline(button_text, b"toggle_process")],
        [Button.inline("🔄 Refresh Stats", b"refresh_stats")],
        [Button.inline("📊 Channel Stats", b"channel_stats")],
        [Button.inline("🗑️ Clear Cache", b"clear_cache")]
    ]
    
    await event.edit(text, buttons=buttons)


@bot.on(events.CallbackQuery(pattern=b"channel_stats"))
async def channel_stats_handler(event):
    """Show detailed channel statistics"""
    if event.sender_id not in config.ADMIN_IDS:
        await event.answer("❌ Unauthorized!", alert=True)
        return
    
    try:
        # Get processed files
        processed_files = load_processed_files()
        
        stats_text = f"""
📊 **DETAILED CHANNEL STATISTICS**

🗂️ **Processed Archives**
• Total Processed: {len(processed_files)}

📁 **Memory Cache**
• Lines in Memory: {len(data_cache):,}
• Cache Status: {'✅ Loaded' if cache_loaded else '❌ Not Loaded'}

📂 **Channels**
• Source: `{config.SOURCE_CHANNEL_ID}`
• Extract: `{extracted_channel_id or 'Not Created'}`
"""
        
        await event.answer(stats_text, alert=True)
        
    except Exception as e:
        await event.answer(f"❌ Error: {e}", alert=True)


@bot.on(events.CallbackQuery(pattern=b"clear_cache"))
async def clear_cache_handler(event):
    """Clear memory cache"""
    global data_cache, cache_loaded
    
    if event.sender_id not in config.ADMIN_IDS:
        await event.answer("❌ Unauthorized!", alert=True)
        return
    
    data_cache = []
    cache_loaded = False
    stats['total_lines'] = 0
    
    await event.answer("🗑️ Cache cleared!", alert=True)
    
    # Update panel
    text = await get_admin_panel_text()
    button_text = "🟢 STOP" if (download_process_running and not download_process_paused) else "🔴 START"
    
    buttons = [
        [Button.inline(button_text, b"toggle_process")],
        [Button.inline("🔄 Refresh Stats", b"refresh_stats")],
        [Button.inline("📊 Channel Stats", b"channel_stats")],
        [Button.inline("🗑️ Clear Cache", b"clear_cache")]
    ]
    
    await event.edit(text, buttons=buttons)


@bot.on(events.NewMessage(pattern='/start'))
async def start_handler(event):
    """Handle /start command"""
    admin_note = "\n\n🔧 Admin? Use /admin for control panel" if event.sender_id in config.ADMIN_IDS else ""
    
    await event.respond(
        "🔍 **Breach Lookup Bot**\n\n"
        "Available commands:\n"
        "/domain <domain> - Domain analysis\n"
        "/email <email> - Email lookup\n"
        "/num <number> - Number lookup\n"
        "/aadhar <aadhar> - Aadhar lookup\n"
        "/pan <pan> - PAN lookup\n"
        "/ps <password> - Password lookup\n"
        "/tg <username> - Telegram lookup\n"
        "/ig <username> - Instagram lookup\n"
        "/fb <username> - Facebook lookup\n"
        "/lk <username> - LinkedIn lookup\n"
        "/x <username> - Twitter/X lookup\n"
        "/yt <channel> - YouTube lookup\n"
        "/wa <number> - WhatsApp lookup\n"
        "/bi <term> - Binance lookup\n"
        "/ot <text> - General text lookup"
        f"{admin_note}"
    )


async def handle_search(event, search_type):
    """Generic search handler"""
    # Get the query
    query = event.text.split(maxsplit=1)
    if len(query) < 2:
        await event.respond(f"❌ Please provide a search term.\nUsage: /{search_type} <query>")
        return
    
    query_text = query[1].strip()
    
    # Check if we can search (either cache loaded OR channel exists)
    if not cache_loaded and not extracted_channel_id:
        await event.respond("⏳ Bot is still setting up. Please wait for first file to be uploaded...")
        return
    
    # Send processing message
    processing_msg = await event.respond("🔍 Searching...")
    
    # Start timing
    start_time = time.time()
    
    # Search
    results = await search_data(query_text, search_type)
    
    # End timing
    elapsed = time.time() - start_time
    
    # Delete processing message
    await processing_msg.delete()
    
    # Prepare response
    response_text = f"✅ **Success**\n\n"
    response_text += f"**Query:** `{query_text}`\n"
    response_text += f"**Time taken:** {elapsed:.2f}s\n"
    response_text += f"**Total hits:** {len(results)}\n"
    
    if len(results) == 0:
        response_text += "\n❌ No results found."
        await event.respond(response_text)
    else:
        # Save results to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"results_{search_type}_{timestamp}.txt"
        filepath = CACHE_DIR / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"Query: {query_text}\n")
            f.write(f"Time taken: {elapsed:.2f}s\n")
            f.write(f"Total hits: {len(results)}\n")
            f.write("-" * 80 + "\n\n")
            for result in results:
                f.write(result + "\n")
        
        # Send response with file
        await event.respond(response_text, file=str(filepath))
        
        # Clean up the temp file
        os.remove(filepath)


# Register all command handlers
@bot.on(events.NewMessage(pattern='/domain'))
async def domain_handler(event):
    await handle_search(event, 'domain')

@bot.on(events.NewMessage(pattern='/email'))
async def email_handler(event):
    await handle_search(event, 'email')

@bot.on(events.NewMessage(pattern='/num'))
async def num_handler(event):
    await handle_search(event, 'num')

@bot.on(events.NewMessage(pattern='/aadhar'))
async def aadhar_handler(event):
    await handle_search(event, 'aadhar')

@bot.on(events.NewMessage(pattern='/pan'))
async def pan_handler(event):
    await handle_search(event, 'pan')

@bot.on(events.NewMessage(pattern='/ps'))
async def ps_handler(event):
    await handle_search(event, 'ps')

@bot.on(events.NewMessage(pattern='/tg'))
async def tg_handler(event):
    await handle_search(event, 'tg')

@bot.on(events.NewMessage(pattern='/ig'))
async def ig_handler(event):
    await handle_search(event, 'ig')

@bot.on(events.NewMessage(pattern='/fb'))
async def fb_handler(event):
    await handle_search(event, 'fb')

@bot.on(events.NewMessage(pattern='/lk'))
async def lk_handler(event):
    await handle_search(event, 'lk')

@bot.on(events.NewMessage(pattern='/x'))
async def x_handler(event):
    await handle_search(event, 'x')

@bot.on(events.NewMessage(pattern='/yt'))
async def yt_handler(event):
    await handle_search(event, 'yt')

@bot.on(events.NewMessage(pattern='/wa'))
async def wa_handler(event):
    await handle_search(event, 'wa')

@bot.on(events.NewMessage(pattern='/bi'))
async def bi_handler(event):
    await handle_search(event, 'bi')

@bot.on(events.NewMessage(pattern='/ot'))
async def ot_handler(event):
    await handle_search(event, 'ot')


async def load_channel_files_into_cache():
    """Load files from extracted channel into memory cache (only if ENABLE_FULL_CACHE is True)"""
    global data_cache, cache_loaded, cache_size_mb
    
    if not extracted_channel_id:
        return
    
    # Check if full caching is enabled
    if not config.ENABLE_FULL_CACHE:
        print("ℹ️ Full memory cache disabled (ENABLE_FULL_CACHE = False)")
        print("ℹ️ Search will load files on-demand from channel (memory efficient)")
        cache_loaded = True
        return
    
    print("📥 Loading channel files into memory cache...")
    print(f"⚠️ Max cache size: {config.MAX_CACHE_SIZE_MB} MB")
    
    try:
        await user_client.start()
        channel = await user_client.get_entity(extracted_channel_id)
        
        # First, count total files
        print("🔍 Counting files in channel...")
        total_files = 0
        async for message in user_client.iter_messages(channel):
            if message.document:
                total_files += 1
        
        print(f"📊 Found {total_files} files to load")
        
        loaded_count = 0
        total_lines = 0
        start_time = time.time()
        
        async for message in user_client.iter_messages(channel):
            if message.document:
                try:
                    file_start = time.time()
                    
                    # Get filename
                    filename = "Unknown"
                    for attr in message.document.attributes:
                        if isinstance(attr, DocumentAttributeFilename):
                            filename = attr.file_name
                            break
                    
                    # Auto-convert binary Excel files to CSV and upload
                    if filename.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                        print(f"   📊 Excel file detected: {filename}")
                        print(f"      🔄 Auto-converting to CSV format...")
                        
                        try:
                            # Download to temp file
                            temp_excel = CACHE_DIR / f"temp_{filename}"
                            await user_client.download_media(message, file=str(temp_excel))
                            
                            # Convert Excel to CSV
                            workbook = openpyxl.load_workbook(temp_excel, read_only=True, data_only=True)
                            converted_lines = []
                            
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                print(f"         Reading sheet: {sheet_name}")
                                
                                for row in sheet.iter_rows(values_only=True):
                                    # Convert row to tab-separated text
                                    row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                                    if row_text.strip():
                                        converted_lines.append(row_text + '\n')
                            
                            workbook.close()
                            
                            # Save as CSV file
                            csv_filename = filename.rsplit('.', 1)[0] + '.csv'
                            csv_path = CACHE_DIR / csv_filename
                            
                            with open(csv_path, 'w', encoding='utf-8', newline='') as f:
                                f.writelines(converted_lines)
                            
                            print(f"      ✅ Converted {len(converted_lines):,} rows to {csv_filename}")
                            
                            # Upload CSV to channel
                            print(f"      📤 Uploading {csv_filename} to channel...")
                            await user_client.send_file(
                                channel,
                                csv_path,
                                caption=f"📊 {csv_filename} (auto-converted from {filename})",
                                force_document=True,
                                part_size_kb=512
                            )
                            print(f"      ✅ Uploaded {csv_filename} to channel")
                            
                            # Add to cache
                            with cache_lock:
                                data_cache.extend(converted_lines)
                            
                            total_lines += len(converted_lines)
                            loaded_count += 1
                            stats['total_lines'] = total_lines
                            stats['loaded_files'] = loaded_count
                            
                            # Check cache size limit
                            cache_size_mb = len(''.join(data_cache).encode('utf-8')) / (1024 * 1024)
                            if cache_size_mb >= config.MAX_CACHE_SIZE_MB:
                                print(f"⚠️ Cache limit reached ({cache_size_mb:.1f} MB)")
                                print(f"⚠️ Stopping cache load. Remaining files will be loaded on-demand.")
                                break
                            
                            # Clean up temp files
                            temp_excel.unlink()
                            csv_path.unlink()
                            
                            # Show progress
                            if loaded_count % 5 == 0 or loaded_count == total_files:
                                elapsed = time.time() - start_time
                                avg_time = elapsed / loaded_count
                                remaining = (total_files - loaded_count) * avg_time
                                print(f"      📊 Progress: {loaded_count}/{total_files} ({loaded_count*100//total_files}%) | "
                                      f"Total: {total_lines:,} lines | ETA: {remaining:.0f}s")
                            
                            continue
                            
                        except Exception as e:
                            print(f"      ❌ Error converting Excel: {e}")
                            print(f"      ⏭️ Skipping this file")
                            loaded_count += 1
                            continue
                    
                    # Reduced logging - only show every 5th file
                    if loaded_count % 5 == 0 or loaded_count == 0:
                        print(f"   📥 [{loaded_count + 1}/{total_files}] Loading: {filename}")
                    
                    # Download file to memory
                    file_bytes = await user_client.download_media(message, file=bytes)
                    
                    file_time = time.time() - file_start
                    file_size_mb = len(file_bytes) / (1024 * 1024)
                    
                    # Try to decode as text with multiple encodings
                    # First check for BOM (Byte Order Mark)
                    file_content = None
                    
                    # Check for UTF-16 BOM
                    if file_bytes[:2] == b'\xff\xfe' or file_bytes[:2] == b'\xfe\xff':
                        try:
                            file_content = file_bytes.decode('utf-16', errors='strict')
                            print(f"      Detected UTF-16 encoding")
                        except:
                            pass
                    
                    # Check for UTF-8 BOM
                    if not file_content and file_bytes[:3] == b'\xef\xbb\xbf':
                        try:
                            file_content = file_bytes[3:].decode('utf-8', errors='strict')
                            print(f"      Detected UTF-8 with BOM")
                        except:
                            pass
                    
                    # Try various encodings
                    if not file_content:
                        for encoding in ['utf-8', 'utf-16-le', 'utf-16-be', 'latin-1', 'cp1252', 'iso-8859-1', 'windows-1252']:
                            try:
                                file_content = file_bytes.decode(encoding, errors='strict')
                                print(f"      Decoded successfully with {encoding}")
                                break
                            except:
                                continue
                    
                    # Fallback
                    if not file_content:
                        file_content = file_bytes.decode('utf-8', errors='ignore')
                        print(f"      Fallback to UTF-8 with errors ignored")
                    
                    # Add lines to cache (keep original lines without stripping for better matching)
                    lines = file_content.splitlines()
                    with cache_lock:
                        for line in lines:
                            if line.strip():  # Only skip completely empty lines
                                data_cache.append(line + '\n')
                    
                    total_lines += len(lines)
                    loaded_count += 1
                    stats['total_lines'] = total_lines
                    stats['loaded_files'] = loaded_count
                    
                    # Check cache size limit
                    cache_size_mb = len(''.join(data_cache).encode('utf-8')) / (1024 * 1024)
                    if cache_size_mb >= config.MAX_CACHE_SIZE_MB:
                        print(f"⚠️ Cache limit reached ({cache_size_mb:.1f} MB)")
                        print(f"⚠️ Stopping cache load. Remaining files will be loaded on-demand.")
                        break
                    
                    # Only show detailed progress every 5th file
                    if loaded_count % 5 == 0 or loaded_count == total_files:
                        elapsed = time.time() - start_time
                        avg_time = elapsed / loaded_count
                        remaining = (total_files - loaded_count) * avg_time
                        
                        print(f"      📊 Progress: {loaded_count}/{total_files} ({loaded_count*100//total_files}%) | "
                              f"Total: {total_lines:,} lines | ETA: {remaining:.0f}s")
                    
                except Exception as e:
                    print(f"   ⚠️ Could not load file: {e}")
                    loaded_count += 1
                    continue
        
        total_time = time.time() - start_time
        cache_loaded = True
        
        print(f"✅ Loaded {loaded_count} files ({len(data_cache):,} lines) into memory in {total_time:.1f}s!")
        await notify_admins(
            f"✅ **Cache Ready!**\n\n"
            f"📊 Loaded: {loaded_count} files\n"
            f"💾 Total lines: {len(data_cache):,}\n"
            f"⏱️ Time taken: {total_time:.1f}s\n\n"
            f"🚀 Ultra-fast search enabled!"
        )
        
    except Exception as e:
        print(f"Error loading channel files: {e}")


async def main():
    """Main function"""
    print("Bot is starting...")
    print("⏸️ Download process is STOPPED by default")
    print("🔧 Admins: Use /admin to START the download process")
    
    # Start user client
    await user_client.start()
    
    # If channel exists, load files into cache immediately on startup
    if extracted_channel_id:
        print("📂 Existing channel found, loading data into memory...")
        await load_channel_files_into_cache()
        print("✅ Memory cache loaded! Search is ready.")
    
    # Don't auto-start downloads - wait for manual start via /admin
    # Create the download task but it will wait for unpause
    download_task = asyncio.create_task(download_and_extract_archives())
    
    print("✅ Bot is ready! Waiting for admin to start downloads...")
    await bot.run_until_disconnected()


if __name__ == '__main__':
    bot.loop.run_until_complete(main())
